from .base import BaseIO
from .pandas import to_dataframe
from collections import OrderedDict
import ast
import pandas as pd
from openpyxl import load_workbook
import re

import sys
import logging
_logging = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG,
                    stream=sys.stdout,
                    format="[%(asctime)s] %(levelname)s:%(name)s:%(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S")


class ExcelIO(BaseIO):
    """
    Reads data from (optionally multiple) worksheets in Microsoft Excel.
    """
    def __init__(self, *args, **kwds):
        super().__init__(*args, **kwds)
        self._options = dict()
        self.set_header(0)
        self.set_sheet_name()

    class Key:
        _sep = '\u0000'
        _fill = ''

        @staticmethod
        def is_container(obj):
            """
            Checks if the object is a container, that is, an iterable,
            but not a string.

            :param obj: Object to be tested.
            :return: True if object is a container, False otherwise.
            """
            if hasattr(obj, '__iter__') and not isinstance(obj, str):
                return True
            return False

        @staticmethod
        def replace_sep(name, sep=' '):
            """
            Replace the separator with a different version.

            :param name, str: Key name to replace.
            :param sep, str: New separator to use in its place.
            :return: Copy of `name` with all instances of the original
                separator replaced. Leading/trailing separators are
                stripped.
            """
            name = re.sub(ExcelIO.Key._sep, sep, name)
            return name.strip(sep)

        @staticmethod
        def to_string(index):
            """
            Converts a MultiIndex into a string.

            :param index: MultiIndex to be converted to a string.
            :return: String version of the index.
            """
            if ExcelIO.Key.is_container(index):
                return ExcelIO.Key._sep.join(str(x) for x in index)
            else:
                return str(index)

        @staticmethod
        def to_index(index, num=-1, squeeze=False):
            """
            Converts a string to an index entry.

            :param index, str: String version of a Index/MultiIndex entry
                generated using `Key.to_string(index)`.
            :param num, int: Minimum levels in the resulting index object.
                If -1 (default), then the resulting tuple is just large
                enough to hold the number of observations. If num is smaller
                than the number of levels, the full complement of levels are
                returned (no information is lost).
            :param squeeze, bool: If the length of the resulting index is
                1, return the contained value instead of a container.
            :return: Index entry
            :rtype: Tuple of entries (if squeeze is False) or scalar value
                (if squeeze is True and only one entry is present).
            """
            levels = tuple(index.split(ExcelIO.Key._sep))
            if len(levels) == 1 and squeeze:
                return levels[0]
            else:
                if num < len(levels):
                    return levels
                else:
                    return levels + (num - len(levels)) * (ExcelIO.Key._fill,)


    def get_sheet_name(self):
        return self._options['sheet_name']


    def set_sheet_name(self, name=None):
        """
        Set which worksheets to read.

        :param name: Name of the worksheet to read. Default: all
            worksheets.
        :return: self
        """
        self._options['sheet_name'] = name
        return self


    def get_header(self, name):
        """
        Returns the number of header rows in the named sheet.

        :param name, str: Name of the worksheet
        :return: int
        """
        if 'header' not in self._options:
            self._options['header'] = dict()
        return self._options['header']\
            .get(name, self._options['default header'])


    def set_header(self, *args, **kwds):
        """
        Set the number of header rows. An unnamed integer argument sets
        the default number of header rows (default: 0). Named keyword
        argument values provide the number of header rows in the worksheet
        whose name matches the key.

        :param args, int: (optional) Default number of header rows.
            Default: 0.
        :param kwds, dict: (optional) Number of header rows in the named
            sheetnames.
        :return: self
        """
        if len(args) > 0:
            self._options['default header'] = int(args[0])
        else:
            try:
                self._options['header'].update(kwds)
            except KeyError:
                self._options['header'] = dict()
                self._options['header'].update(kwds)
        return self


    def read_worksheet(self, sheet):
        values = sheet.values
        num = self.get_header(sheet.title)
        _logging.debug(f"{num} header row(s) in {sheet.title}")
        columns = [[('' if v is None else str(v))
                    for v in next(values)[0:]]
                   for _ in range(num)]
        if len(columns) == 0:
            columns = None
        else:
            columns = list(zip(*columns))
        _logging.debug(f"Columns: {columns}")
        return pd.DataFrame(values, columns=columns)


    def read_workbook(self, workbook):
        sheets = self._options['sheet_name']
        if sheets is None:
            sheets = workbook.sheetnames
        results = OrderedDict()
        for sheet in sheets:
            results[sheet] = self.read_worksheet(workbook[sheet])
        return results


    def read_excel(self, fname):
        return self.read_workbook(load_workbook(fname))


    def load(self, fobj, *args, **kwds):
        """
        Loads node data from the specified Excel workbook.

        :param fobj: File to be read.
        :type fobj: valid io object to pandas.read_excel.
        :param sheet_name: Sheetnames to be read. Default: all.
        :type sheet_name: str, iterable of str, or None (all, default)
        :return: Nodes read from the Excel workbook.
        :rtype: list of Nodes
        """
        def convert(obj):
            """
            Pandas `read_excel` has the `converters` option that allows the
            user to specify converters for each column. The value of this
            option must be a dictionary that maps the column name or index
            of the column to a function that accepts the cell data (as a
            string) and returns the converted value.

            If python-style lists, tuples, dictionaries, etc. are stored
            in an Excel cell, Converter calls ast.literal_eval as a conversion
            for each cell regardless of the key (column name/column index).
            """
            try:
                return ast.literal_eval(obj)
            except (ValueError, SyntaxError):
                return obj

        # set defaults, e.g. sheet_name --> None
        self.set_sheet_name(kwds.get('sheet_name', None))
        if 'header' in kwds:
            self.set_header(kwds['header'])
        # read the file
        df = self.read_excel(fobj)
        # convert each element to handle lists, tuples, etc.
        _logging.debug(f"Reading {fobj}...")
        for sheet_name in df:
            _logging.debug(f"Reading {sheet_name}...")
            df[sheet_name] = df[sheet_name].applymap(convert)
            _logging.debug(f"Read {len(df[sheet_name])} entries "
                           f"from {sheet_name}.")
        _logging.debug(f"Read {len(df)} sheets from {fobj}.")
        # convert each entry into a node
        _logging.debug(f"Creating nodes from {fobj}...")
        try:
            if not isinstance(df, dict):
                df = {None: df}
            nodes = [self[key](**{ExcelIO.Key.to_string(k):v
                                  for (k, v) in row.to_dict().items()})
                     for key in df.keys()
                     for (index, row) in df[key].iterrows()]
        except:
            msg = '\n'.join([
                '{:<15s} {:>15s}'.format(ExcelIO.Key.to_string(k),
                                          str(v))
                for key in df.keys()
                for (index, row) in df[key].iterrows()
                for (k, v) in row.to_dict().items()])
            # _logging.debug(msg)
            _logging.debug('\n'.join([str(d.columns.to_list())
                                      for d in df.values()]))
            raise
        _logging.debug(f"Created {len(nodes)} nodes from {fobj}.")
        # done
        return nodes

    def dump(self, fobj, *args, **kwds):
        """
        Dumps args (lists of nodes) to a file-like object.

        :param fobj: File-like object to which the nodes are written in Excel
            format.
        :type fobj: File-like object (str or file)
        :param args: Lists of nodes that are to be written to the file object.
        :type args: list of nodes
        :param kwds: Keywords passed to `ExcelIO.Key.to_index`.
        :return: None
        """
        nodes = []
        for arg in args:
            nodes.extend(arg)
        df = to_dataframe(nodes)
        df.columns = pd.MultiIndex.from_tuples(
            [ExcelIO.Key.to_index(name, **kwds) for name in df.columns])
        df.to_excel(fobj, index=False)
